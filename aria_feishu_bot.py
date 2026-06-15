"""
aria_feishu_bot.py — Feishu (Lark) 多模态 AI 机器人
======================================================
OpenClaw 同款设计：任意输入（文字/语音/图片/文件）→ Aria AI → 卡片回复

两种运行模式：
  1. 嵌入 FastAPI（由 feishu_routes.py 调用）
  2. 独立运行  — python3 aria_feishu_bot.py [port]

飞书端配置：
  1. 飞书开发者后台 → 创建自建应用
  2. 事件订阅 → Request URL: http://<host>/api/v1/feishu/event
  3. 权限：im:message / im:message:send_as_bot
  4. ~/.aria/.env 填写：
       FEISHU_APP_ID=cli_xxx        FEISHU_APP_SECRET=xxx
       ANTHROPIC_API_KEY=xxx        # 图片理解 / LLM
       OPENAI_API_KEY=xxx           # Whisper 语音转文字（可选）
       FEISHU_ALLOWED_USER_IDS=uid1,uid2   # 留空=不限制

支持的消息类型：
  📝 文字（非命令）→ Aria LLM 自然语言回答
  🎤 语音          → Whisper 转文字 → Aria LLM
  🖼️  图片          → 视觉 LLM 分析（Claude / GPT-4V）
  📄 文件          → 自动解析 PDF/Excel/代码 → Aria LLM 总结

结构化命令（/command）：
  /price AAPL           /brief          /screen
  /report NVDA          /run /price TSLA   (调用 aria CLI -p 模式)
  /alert add SYM cond v  /alerts         /status   /help
"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import hmac
import json
import logging
import os
import re
import sys
import tempfile
import time
from pathlib import Path
from typing import Any, Dict, Optional

logger = logging.getLogger(__name__)

# Strip ANSI escape codes from aria CLI output
_ANSI_RE = re.compile(r"\x1b(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])")

# Lines that are tool/UI artifacts and must be stripped from bot replies
_BOT_NOISE_RE = re.compile(
    r"^\s*(?:"
    # ── diff / table lines (ASCII pipe AND Unicode box-drawing │ U+2502) ────
    r"[│|][+\- \d]"                    # │298  │+ code  |- old  | context
    r"|[│|]\s*$"                        # │      │  (empty cell borders)
    r"|[┌┐└┘├┤┬┴┼─╌╍╴╶╷╸╹]"          # box corners / connectors
    # ── timing artifacts ────────────────────────────────────────────────────
    r"|└\s*\d+[\.,]\d+s?\b"
    r"|[└─]{1,3}\s*\d+[\.,]\d+\s*s"
    # ── tool call / result bullets ──────────────────────────────────────────
    r"|  [●└■▸]"
    r"|  L \d"
    # ── permission / confirmation dialog ────────────────────────────────────
    r"|[›❯>]\s*\d+\."                  # › 1. Yes  ❯ 1. Yes  > 1. Yes
    r"|\d+\.\s+Yes"                    # 1. Yes / 2. Yes, allow all
    r"|\d+\.\s+No"                     # 3. No
    r"|Enter number"                   # "Enter number (or Enter to keep current):"
    r"|Cancelled"
    # ── leftover Rich markup tags ────────────────────────────────────────────
    r"|\[/?(?:cyan|dim|bold|red|green|yellow|blue|magenta|white|grey|reset)\]"
    # ── horizontal rules ────────────────────────────────────────────────────
    r"|\s*[━─═]{4,}\s*$"
    r")"
)

# Inline timing / markup to strip from within a line
_INLINE_TIMING_RE = re.compile(r"\s*[└─]{1,2}\s*\d+[\.,]\d+\s*s\b")
_INLINE_RICH_TAG_RE = re.compile(r"\[/?(?:cyan|dim|bold|red|green|yellow|blue|magenta|white|grey|reset)\]")

_ARIA_CODE_DIR = Path(__file__).parent

# ── Feishu API endpoints ───────────────────────────────────────────────────────

_FEISHU_API = "https://open.feishu.cn/open-apis"

# ── Token cache (tenant_access_token, expires ~2h) ────────────────────────────
_token_cache: dict[str, Any] = {"token": None, "expires_at": 0}


async def _get_access_token() -> Optional[str]:
    """Fetch/cache tenant_access_token via app credentials."""
    app_id     = os.environ.get("FEISHU_APP_ID", "")
    app_secret = os.environ.get("FEISHU_APP_SECRET", "")
    if not app_id or not app_secret:
        return None

    now = time.time()
    if _token_cache["token"] and _token_cache["expires_at"] > now + 60:
        return _token_cache["token"]

    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                f"{_FEISHU_API}/auth/v3/tenant_access_token/internal",
                json={"app_id": app_id, "app_secret": app_secret},
            )
            data = resp.json()
            token = data.get("tenant_access_token")
            expire = int(data.get("expire", 7200))
            _token_cache["token"] = token
            _token_cache["expires_at"] = now + expire
            return token
    except Exception as exc:
        logger.warning("Feishu token fetch failed: %s", exc)
        return None


# ── Send message helpers ───────────────────────────────────────────────────────

async def _feishu_post(url: str, token: str, payload: dict) -> Optional[dict]:
    """POST to Feishu API; log the response code on error."""
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(
                url,
                headers={"Authorization": f"Bearer {token}",
                         "Content-Type": "application/json; charset=utf-8"},
                json=payload,
            )
        data = resp.json()
        code = data.get("code", 0)
        if code != 0:
            logger.error("Feishu API error %s: %s  url=%s  msg_id in payload=%s",
                         code, data.get("msg", ""), url.split("/")[-3:],
                         payload.get("receive_id", "—"))
        else:
            logger.debug("Feishu API ok: %s", url.split("/")[-2:])
        return data
    except Exception as exc:
        logger.warning("Feishu POST failed: %s", exc)
        return None


async def reply_text(message_id: str, text: str) -> None:
    """Reply to a Feishu message with plain text (auto-truncated at 3000 chars)."""
    if not message_id:
        logger.error("reply_text: empty message_id — cannot reply")
        return
    token = await _get_access_token()
    if not token:
        logger.error("reply_text: no access token")
        return
    logger.info("reply_text → message_id=%s len=%d", message_id, len(text))
    await _feishu_post(
        f"{_FEISHU_API}/im/v1/messages/{message_id}/reply",
        token,
        {"msg_type": "text", "content": json.dumps({"text": text[:3000]})},
    )


async def reply_card(message_id: str, title: str, body: str,
                     color: str = "blue", footer: str = "") -> None:
    """Reply with an interactive card (title + Markdown body)."""
    if not message_id:
        logger.error("reply_card: empty message_id — cannot reply")
        return
    token = await _get_access_token()
    if not token:
        await reply_text(message_id, f"【{title}】\n{body}")
        return
    elements = _build_card_elements(body, footer)
    card = {
        "config": {"wide_screen_mode": True},
        "header": {"title": {"tag": "plain_text", "content": title}, "template": color},
        "elements": elements,
    }
    logger.info("reply_card → message_id=%s title=%s", message_id, title[:40])
    result = await _feishu_post(
        f"{_FEISHU_API}/im/v1/messages/{message_id}/reply",
        token,
        {"msg_type": "interactive", "content": json.dumps(card)},
    )
    # If card failed, fall back to plain text
    if result and result.get("code") != 0:
        logger.info("reply_card: card failed (code %s), falling back to text", result.get("code"))
        await reply_text(message_id, f"【{title}】\n{body}")


async def reply_or_send(message_id: str, chat_id: str,
                        title: str, body: str,
                        color: str = "blue", footer: str = "") -> None:
    """Try reply by message_id first; if that fails, send a new message to chat_id."""
    if message_id:
        result = await _reply_card_raw(message_id, title, body, color, footer)
        if result is not None and result.get("code", 0) == 0:
            return
        logger.warning("reply failed (code=%s), falling back to send_card_to_chat",
                       result.get("code") if result else "no response")
    if chat_id:
        await send_card_to_chat(chat_id, title, body, color)
    else:
        logger.error("reply_or_send: both message_id and chat_id empty, cannot send")


def _build_card_elements(body: str, footer: str = "") -> list:
    """Split body into visual sections for richer card layout."""
    import re as _re
    # Split on markdown `---` dividers or `##`/`###` section headers
    _section_re = _re.compile(r'(?m)^[-─]{3,}\s*$')
    raw_sections = _section_re.split(body.strip())
    elements: list = []
    for i, sec in enumerate(raw_sections):
        sec = sec.strip()
        if not sec:
            continue
        # Detect if section starts with a ## header and peel it off
        _hdr_m = _re.match(r'^#{1,3}\s+(.+)\n', sec)
        if _hdr_m:
            hdr_text = _hdr_m.group(1).strip()
            sec_body = sec[_hdr_m.end():].strip()
            elements.append({"tag": "markdown", "content": f"**{hdr_text}**"})
            if sec_body:
                elements.append({"tag": "div", "text": {"tag": "lark_md", "content": sec_body[:900]}})
        else:
            elements.append({"tag": "div", "text": {"tag": "lark_md", "content": sec[:900]}})
        if i < len(raw_sections) - 1:
            elements.append({"tag": "hr"})

    if not elements:
        elements = [{"tag": "div", "text": {"tag": "lark_md", "content": body[:2000]}}]

    if footer:
        elements += [{"tag": "hr"}, {"tag": "note", "elements": [
            {"tag": "plain_text", "content": footer}
        ]}]
    return elements


async def _reply_card_raw(message_id: str, title: str, body: str,
                          color: str = "blue", footer: str = "") -> Optional[dict]:
    """Reply with a card; return raw API response dict."""
    token = await _get_access_token()
    if not token:
        return None
    elements = _build_card_elements(body, footer)
    card = {
        "config": {"wide_screen_mode": True},
        "header": {"title": {"tag": "plain_text", "content": title}, "template": color},
        "elements": elements,
    }
    return await _feishu_post(
        f"{_FEISHU_API}/im/v1/messages/{message_id}/reply",
        token,
        {"msg_type": "interactive", "content": json.dumps(card)},
    )


async def send_card_to_chat(chat_id: str, title: str, body: str,
                            color: str = "blue", receive_id_type: str = "chat_id") -> None:
    """Send a new card message to a chat (group or user)."""
    token = await _get_access_token()
    if not token:
        return
    elements = _build_card_elements(body)
    card = {
        "config": {"wide_screen_mode": True},
        "header": {"title": {"tag": "plain_text", "content": title}, "template": color},
        "elements": elements,
    }
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15) as client:
            await client.post(
                f"{_FEISHU_API}/im/v1/messages?receive_id_type={receive_id_type}",
                headers={"Authorization": f"Bearer {token}"},
                json={
                    "receive_id": chat_id,
                    "msg_type":   "interactive",
                    "content":    json.dumps({"card": card}),
                },
            )
    except Exception as exc:
        logger.warning("send_card_to_chat failed: %s", exc)


# ── Command router ─────────────────────────────────────────────────────────────

async def _handle_command(cmd: str, message_id: str, sender_id: str, chat_id: str = "") -> None:
    """Parse a command string and reply with structured card."""
    parts = cmd.strip().split()
    if not parts:
        return
    verb = parts[0].lstrip("/").lower()

    if verb == "help":
        body = (
            "**💬 直接发消息** — 自然语言提问，Aria AI 直接回答\n"
            "**🎤 语音消息** — 自动转文字后 AI 分析\n"
            "**🖼️ 图片** — 自动识别图表/截图内容\n"
            "**📄 文件** — PDF/Excel/Word/代码 自动解析后 AI 总结\n\n"
            "**结构化命令：**\n"
            "`/price <symbol>` — 实时价格（支持 A 股 6 位代码）\n"
            "`/brief` — 今日晨报摘要\n"
            "`/screen` — 涨停预测 Top10\n"
            "`/report <symbol>` — 个股研报（异步推送）\n"
            "`/team <symbol>` — 🤖 多Agent研究（宏观+基本面+技术+风控）\n"
            "`/football predict Arsenal vs Chelsea pl` — ⚽ 足球比赛预测\n"
            "`/football standings pl` — 联赛积分榜（pl/bl/ll/sa/cl）\n"
            "`/run <aria命令>` — 执行任意 Aria 命令，如 `/run /corr AAPL TSLA`\n"
            "`/alert add <symbol> <cond> <value>` — 添加价格预警\n"
            "　　条件: `price_above` `price_below` `pct_change_above` `pct_change_below`\n"
            "`/alerts` — 查看所有预警\n"
            "`/status` — Daemon 运行状态\n"
            "`/help` — 显示此帮助"
        )
        await reply_card(message_id, "📖 Aria 帮助", body, "blue")

    elif verb == "price":
        symbol = parts[1].upper() if len(parts) > 1 else ""
        if not symbol:
            await reply_text(message_id, "用法: /price <symbol>，例如 /price AAPL 或 /price 600036")
            return
        await reply_card(message_id, f"🔄 查询 {symbol}…", "正在获取实时行情，请稍候…", "blue")
        try:
            price, prev = await _fetch_price_feishu(symbol)
            if price is None:
                await reply_card(message_id, f"❌ {symbol}", "获取行情失败，请检查代码是否正确", "red")
                return
            pct = f"{(price - prev) / prev * 100:+.2f}%" if prev else "N/A"
            color = "red" if prev and price > prev else "green" if prev and price < prev else "blue"
            body = (
                f"**当前价** ¥{price:.3f}\n"
                f"**涨跌幅** {pct}\n"
                f"**昨收**  ¥{prev:.3f}" if prev else f"**当前价** {price:.4f}"
            )
            await reply_card(message_id, f"{'📈' if prev and price >= prev else '📉'} {symbol}", body, color)
        except Exception as exc:
            await reply_card(message_id, f"❌ {symbol}", f"查询失败: {exc}", "red")

    elif verb == "brief":
        await reply_card(message_id, "⏳ 生成晨报…", "正在获取市场数据，请稍候…", "blue")
        try:
            from aria_daemon import _run_morning_brief
            brief = await _run_morning_brief()
            await reply_card(message_id, "📊 Aria 晨报", brief[:2000], "green",
                             footer="Aria Code · 实时市场分析")
        except Exception as exc:
            await reply_card(message_id, "❌ 晨报生成失败", str(exc)[:300], "red")

    elif verb == "screen":
        await reply_card(message_id, "⏳ 筛选中…", "正在扫描 A 股涨停预测，请稍候…", "blue")
        try:
            from aria_daemon import _run_screener
            result = await _run_screener()
            await reply_card(message_id, "🔍 涨停预测 Top10", result[:2000], "turquoise")
        except Exception as exc:
            await reply_card(message_id, "❌ 筛选失败", str(exc)[:300], "red")

    elif verb == "report":
        symbol = parts[1].upper() if len(parts) > 1 else ""
        if not symbol:
            await reply_text(message_id, "用法: /report <symbol>，例如 /report NVDA")
            return
        await reply_card(message_id, f"⏳ 生成 {symbol} 研报…",
                         "正在进行多维度分析，通常需要 30-60 秒，完成后推送结果。", "blue")
        asyncio.create_task(_async_report(symbol, message_id))

    elif verb == "alert":
        if len(parts) < 2:
            await reply_text(message_id, "用法: /alert add <symbol> <cond> <value>")
            return
        sub = parts[1].lower()
        if sub == "add" and len(parts) >= 5:
            sym, cond, val = parts[2].upper(), parts[3], parts[4]
            valid_conds = {"price_above", "price_below", "pct_change_above", "pct_change_below"}
            if cond not in valid_conds:
                await reply_card(message_id, "❌ 无效条件",
                                 f"支持的条件：\n" + "\n".join(f"• `{c}`" for c in sorted(valid_conds)), "red")
                return
            import sqlite3
            from aria_daemon import _DB_PATH
            with sqlite3.connect(_DB_PATH) as conn:
                conn.execute(
                    "INSERT INTO alerts(id,symbol,condition,value,message,active) VALUES(?,?,?,?,?,1)",
                    (f"fs_{int(time.time())}_{sym}", sym, cond, float(val),
                     f"{sym} {cond.replace('_',' ')} {val}", )
                )
                conn.commit()
            await reply_card(message_id, "✅ 预警已添加",
                             f"**{sym}** 当 {cond.replace('_',' ')} `{val}` 时触发通知", "green")
        else:
            await reply_text(message_id, "用法: /alert add <symbol> <cond> <value>")

    elif verb == "alerts":
        import sqlite3
        from aria_daemon import _DB_PATH
        rows = sqlite3.connect(_DB_PATH).execute(
            "SELECT symbol,condition,value FROM alerts WHERE active=1 ORDER BY created_at DESC LIMIT 20"
        ).fetchall()
        if not rows:
            await reply_card(message_id, "📋 当前预警", "暂无活跃预警", "blue")
        else:
            lines = "\n".join(f"• **{r[0]}** {r[1].replace('_',' ')} `{r[2]}`" for r in rows)
            await reply_card(message_id, f"📋 活跃预警（{len(rows)} 条）", lines, "blue")

    elif verb == "status":
        import sqlite3
        from aria_daemon import _DB_PATH, _PID_FILE
        pid_alive = _PID_FILE.exists()
        conn = sqlite3.connect(_DB_PATH)
        alert_count = conn.execute("SELECT COUNT(*) FROM alerts WHERE active=1").fetchone()[0]
        sched_count = conn.execute("SELECT COUNT(*) FROM schedules WHERE active=1").fetchone()[0]
        job_pending = conn.execute("SELECT COUNT(*) FROM webhook_jobs WHERE status='pending'").fetchone()[0]
        body = (
            f"**Daemon** {'🟢 运行中' if pid_alive else '🔴 未运行'}\n"
            f"**活跃预警** {alert_count} 条\n"
            f"**定时任务** {sched_count} 条\n"
            f"**待处理 Jobs** {job_pending} 个"
        )
        await reply_card(message_id, "📡 Aria Daemon 状态", body,
                         "green" if pid_alive else "red")
    elif verb == "football":
        sub = parts[1].lower() if len(parts) > 1 else ""
        rest = " ".join(parts[2:])
        if sub == "predict" and " vs " in rest.lower():
            await reply_card(message_id, f"⚽ 预测中…", f"> {rest}", "blue")
            asyncio.create_task(_handle_football_predict(rest, message_id))
        elif sub == "standings":
            league = rest.strip() or "pl"
            await reply_card(message_id, f"⚽ 获取积分榜…", f"联赛: {league.upper()}", "blue")
            asyncio.create_task(_handle_football_standings(league, message_id))
        else:
            # Natural language after /football (e.g. "/football 预测加拿大跟波黑")
            # → treat as NL query with football intent, route to LLM
            nl_text = f"{sub} {rest}".strip() if rest else sub
            _is_chinese = any('一' <= c <= '鿿' for c in nl_text)
            _is_predict_kw = any(k in nl_text.lower() for k in (
                "predict", "preview", "who wins", "who will", "预测", "谁赢", "比分", "胜率"
            ))
            if _is_chinese or _is_predict_kw:
                asyncio.create_task(_handle_nl_query(nl_text, message_id, chat_id))
            else:
                await reply_card(
                    message_id, "⚽ /football 命令",
                    f"未识别子命令: `{sub}`\n\n**用法:**\n"
                    "- `/football predict Arsenal vs Chelsea pl` — 预测比赛\n"
                    "- `/football standings pl` — 积分榜\n"
                    "- 或直接用自然语言提问，例如：「预测加拿大跟波黑的比分」",
                    "yellow"
                )

    elif verb in ("team", "analyze"):
        # /team <symbol> [--full]  — multi-agent research team
        sym_parts = [p for p in parts[1:] if not p.startswith("-")]
        flags     = [p for p in parts[1:] if p.startswith("-")]
        symbol    = sym_parts[0].upper() if sym_parts else ""
        if not symbol:
            await reply_card(message_id, "❓ 用法",
                             "`/team <symbol>` — 多Agent研究\n"
                             "`/team AAPL --full` — 完整7-agent模式\n"
                             "例: `/team NVDA`  `/team 600519`", "blue")
            return
        flag_str = " " + " ".join(flags) if flags else ""
        cmd = f"/team {symbol}{flag_str}"
        await reply_card(message_id, f"🤖 多Agent分析 {symbol}…",
                         f"正在启动4-agent并行分析，请稍候（约15-30s）…", "blue")
        asyncio.create_task(_async_run_aria(cmd, message_id))

    elif verb == "run":
        # /run <aria-command>  e.g. /run /price AAPL  /run /corr AAPL TSLA NVDA
        sub_cmd = " ".join(parts[1:]).strip() if len(parts) > 1 else ""
        if not sub_cmd:
            await reply_text(message_id, "用法: /run <aria命令>，例如 `/run /price AAPL`")
            return
        await reply_card(message_id, f"⚙️ 执行: {sub_cmd[:60]}", "正在运行，请稍候…", "blue")
        asyncio.create_task(_async_run_aria(sub_cmd, message_id))

    else:
        await reply_card(message_id, "❓ 未知命令",
                         f"不认识 `/{verb}`，发送 `/help` 查看全部命令", "yellow")


async def _handle_football_predict(match_str: str, message_id: str) -> None:
    """Handle /football predict <home> vs <away> [league] from Feishu."""
    import re
    m = re.match(r"(.+?)\s+vs\s+(.+?)(?:\s+(\w+))?$", match_str, re.IGNORECASE)
    if not m:
        await reply_card(message_id, "❌ 格式错误", "用法: `/football predict Arsenal vs Chelsea pl`", "red")
        return
    home_raw, away_raw, league = m.group(1).strip(), m.group(2).strip(), (m.group(3) or "pl")
    try:
        from football_data_client import _CN_TEAM_MAP, _FIFA_RATINGS, predict_match, predict_wc_match

        # Translate Chinese team names → English for model lookup
        home_en = _CN_TEAM_MAP.get(home_raw, home_raw)
        away_en = _CN_TEAM_MAP.get(away_raw, away_raw)
        home_low = home_en.lower().strip()
        away_low = away_en.lower().strip()

        # Use WC/national team model when both teams are in FIFA ratings table
        if home_low in _FIFA_RATINGS and away_low in _FIFA_RATINGS:
            pred = predict_wc_match(home_en, away_en, neutral_venue=True)
            home_label = pred.get("home_name_cn") or home_raw
            away_label = pred.get("away_name_cn") or away_raw
            ranking_note = (
                f"FIFA 排名: #{pred.get('home_ranking','?')} vs #{pred.get('away_ranking','?')}"
            )
            most_likely = pred["top_scorelines"][0]["score"] if pred.get("top_scorelines") else "?"
        else:
            pred = predict_match(home_en, away_en, league)
            home_label, away_label = home_raw, away_raw
            ranking_note = f"联赛: {league.upper()}"
            most_likely = pred.get("most_likely_score") or (
                pred["top_scorelines"][0]["score"] if pred.get("top_scorelines") else "?"
            )

        top = "\n".join(
            f"  {s['score']} — {s['prob']}%" for s in (pred.get("top_scorelines") or [])[:3]
        )
        body = (
            f"**{home_label}** vs **{away_label}**\n"
            f"*{ranking_note}*\n\n"
            f"---\n"
            f"🏆 主队胜: **{pred['home_win']:.0%}**  "
            f"平局: {pred['draw']:.0%}  "
            f"客队胜: {pred['away_win']:.0%}\n\n"
            f"⚽ 预期进球: {pred['lambda_home']:.1f} – {pred['lambda_away']:.1f}\n"
            f"📊 最可能比分: **{most_likely}**\n"
            f"🎯 高概率比分:\n{top}\n\n"
            f"双方均进球: {pred['btts']:.0%}\n"
            f"*泊松模型量化预测，仅供参考*"
        )
        color = (
            "green" if pred["home_win"] > pred["away_win"] + 0.1
            else "red" if pred["away_win"] > pred["home_win"] + 0.1
            else "yellow"
        )
        await reply_card(message_id, "⚽ 赛事预测", body, color)
    except Exception as exc:
        await reply_card(message_id, "❌ 预测失败", str(exc)[:300], "red")


async def _handle_football_standings(league: str, message_id: str) -> None:
    """Handle /football standings from Feishu."""
    try:
        from football_data_client import get_standings
        data = get_standings(league)
        if not data:
            await reply_card(message_id, "❌ 无法获取数据",
                             "请检查联赛代码或设置 FOOTBALL_DATA_API_KEY", "red")
            return
        rows = data["table"][:10]
        lines = [f"**{data['league_name']}**\n"]
        for r in rows:
            form = r.get("form", "") or ""
            lines.append(f"{r['pos']:2}. {r['team'][:18]:18} {r['pts']}分  {r['w']}W{r['d']}D{r['l']}L  {form[:5]}")
        await reply_card(message_id, "📊 积分榜 TOP10", "\n".join(lines), "blue")
    except Exception as exc:
        await reply_card(message_id, "❌ 获取失败", str(exc)[:300], "red")


async def _fetch_price_feishu(symbol: str):
    """Thin wrapper around aria_daemon._fetch_price for use inside bot."""
    try:
        from aria_daemon import _fetch_price
        return await _fetch_price(symbol)
    except ImportError:
        # Fallback: inline yfinance
        import yfinance as yf
        yfn = (symbol + (".SS" if symbol.startswith(("6", "5")) else ".SZ")
               if symbol.isdigit() and len(symbol) == 6 else symbol)
        info = yf.Ticker(yfn).fast_info
        p = getattr(info, "last_price", None)
        pc = getattr(info, "previous_close", None)
        return (float(p) if p else None, float(pc) if pc else None)


async def _async_report(symbol: str, message_id: str) -> None:
    """Background task: generate report and reply when done."""
    try:
        from aria_daemon import _run_report
        result = await _run_report(symbol)
        await reply_card(message_id, f"📄 {symbol} 研报完成", result[:2000], "turquoise",
                         footer="Aria Code · 多智能体分析")
    except Exception as exc:
        await reply_card(message_id, f"❌ {symbol} 研报失败", str(exc)[:300], "red")


async def _async_run_aria(cmd: str, message_id: str) -> None:
    """Background task: run aria CLI command and reply with result."""
    result = await _query_aria_llm(cmd, timeout=120)
    color = "red" if result.startswith("❌") or result.startswith("⏱️") else "turquoise"
    await reply_card(message_id, f"✅ Aria 执行完成", result[:2000], color,
                     footer=f"命令: {cmd[:80]}")


# ── Multimodal helpers ────────────────────────────────────────────────────────

def _is_allowed_user(user_id: str) -> bool:
    """Check FEISHU_ALLOWED_USER_IDS allowlist (empty = allow all)."""
    raw = os.environ.get("FEISHU_ALLOWED_USER_IDS", "").strip()
    if not raw:
        return True
    return user_id in {u.strip() for u in raw.split(",") if u.strip()}


async def _query_aria_direct(text: str, timeout: int = 90) -> str:
    """
    Query the LLM directly via providers/llm/registry.py — no subprocess, no tool use.
    Used for conversational NL queries where we want a clean text answer.
    Falls back to _query_aria_llm on ImportError.
    """
    try:
        import sys as _sys
        _sys.path.insert(0, str(_ARIA_CODE_DIR))
        from providers.llm.registry import stream_cloud_fallback
        import asyncio as _aio

        collected: list[str] = []
        result = await _aio.wait_for(
            stream_cloud_fallback(text, history=[], on_token=collected.append),
            timeout=timeout,
        )
        if result.get("success") and collected:
            return "".join(collected).strip()
        if result.get("response"):
            return result["response"].strip()
        return "❌ LLM 返回空响应，请检查 API Key 配置。"
    except ImportError:
        pass  # fall through to subprocess
    except Exception as _exc:
        logger.warning("_query_aria_direct failed: %s", _exc)
    return await _query_aria_llm(text, timeout=timeout)


async def _query_aria_llm(text: str, timeout: int = 120) -> str:
    """
    Run a natural language query through aria-code's LLM via CLI -p mode.
    Returns plain text output (ANSI stripped, tool noise filtered).
    Use _query_aria_direct() for conversational queries — this is for
    slash commands (/brief, /team, etc.) that need the full CLI context.
    """
    aria_cli = _ARIA_CODE_DIR / "aria_cli.py"
    if not aria_cli.exists():
        return "❌ aria_cli.py 未找到，请检查 ARIA_CODE_DIR 配置。"
    try:
        # ARIA_BOT_MODE=1: auto-approves tools + suppresses visual diffs in aria_cli
        bot_env = {**os.environ, "ARIA_BOT_MODE": "1"}
        proc = await asyncio.create_subprocess_exec(
            sys.executable, str(aria_cli), "-p", text,
            stdin=asyncio.subprocess.DEVNULL,   # no interactive prompts
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=str(_ARIA_CODE_DIR),
            env=bot_env,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        raw = stdout.decode("utf-8", errors="replace")
        # Strip ANSI codes, filter noise lines, collapse blank runs
        clean_lines = []
        blank_run = 0
        for line in _ANSI_RE.sub("", raw).splitlines():
            if _BOT_NOISE_RE.match(line):
                continue
            line = _INLINE_TIMING_RE.sub("", line)
            line = _INLINE_RICH_TAG_RE.sub("", line)
            if not line.strip():
                blank_run += 1
                if blank_run > 1:   # collapse consecutive blank lines to one
                    continue
            else:
                blank_run = 0
            clean_lines.append(line)
        clean = "\n".join(clean_lines).strip()
        return clean[:3000] if clean else (stderr.decode()[:500] or "（无输出）")
    except asyncio.TimeoutError:
        try: proc.kill()
        except Exception: pass
        return f"⏱️ 查询超时（>{timeout}s），请简化问题后重试。"
    except Exception as exc:
        return f"❌ aria LLM 调用失败: {exc}"


async def _download_feishu_resource(message_id: str, resource_key: str,
                                    rtype: str = "file") -> Optional[bytes]:
    """Download image / audio / file from Feishu message."""
    token = await _get_access_token()
    if not token:
        return None
    try:
        import httpx
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.get(
                f"{_FEISHU_API}/im/v1/messages/{message_id}/resources/{resource_key}",
                params={"type": rtype},
                headers={"Authorization": f"Bearer {token}"},
            )
            if resp.status_code == 200:
                return resp.content
    except Exception as exc:
        logger.warning("_download_feishu_resource failed: %s", exc)
    return None


async def _transcribe_voice(audio_bytes: bytes) -> str:
    """Speech-to-text: tries OpenAI Whisper API, then local faster-whisper."""
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if api_key:
        try:
            import httpx
            with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as f:
                f.write(audio_bytes)
                tmp = f.name
            with open(tmp, "rb") as audio_file:
                async with httpx.AsyncClient(timeout=60) as client:
                    resp = await client.post(
                        "https://api.openai.com/v1/audio/transcriptions",
                        headers={"Authorization": f"Bearer {api_key}"},
                        data={"model": "whisper-1"},
                        files={"file": ("voice.ogg", audio_file, "audio/ogg")},
                    )
                    data = resp.json()
                    return data.get("text", "（识别结果为空）")
        except Exception as exc:
            logger.warning("OpenAI Whisper failed: %s", exc)

    # Fallback: local faster-whisper / openai-whisper
    try:
        from faster_whisper import WhisperModel
        model = WhisperModel("base", device="cpu", compute_type="int8")
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as f:
            f.write(audio_bytes)
            tmp = f.name
        segments, _ = model.transcribe(tmp, language="zh")
        return " ".join(seg.text for seg in segments).strip() or "（识别结果为空）"
    except ImportError:
        pass

    return "❌ 语音转文字需要配置 OPENAI_API_KEY，或安装 faster-whisper (`pip install faster-whisper`)"


async def _analyze_image(image_bytes: bytes, caption: str = "") -> str:
    """Analyze image via Claude vision (ANTHROPIC_API_KEY) or GPT-4V (OPENAI_API_KEY)."""
    b64 = base64.b64encode(image_bytes).decode()
    prompt = caption or "请详细分析这张图片的内容。如果是图表、K线图或截图，请做专业解读并给出结论。"

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if anthropic_key:
        try:
            import httpx
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": anthropic_key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json={
                        "model": "claude-sonnet-4-6",
                        "max_tokens": 1500,
                        "messages": [{
                            "role": "user",
                            "content": [
                                {"type": "image", "source": {
                                    "type": "base64", "media_type": "image/jpeg", "data": b64}},
                                {"type": "text", "text": prompt},
                            ],
                        }],
                    },
                )
                data = resp.json()
                return data["content"][0]["text"]
        except Exception as exc:
            logger.warning("Claude vision failed: %s", exc)

    openai_key = os.environ.get("OPENAI_API_KEY", "")
    if openai_key:
        try:
            import httpx
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {openai_key}"},
                    json={
                        "model": "gpt-4o",
                        "max_tokens": 1500,
                        "messages": [{
                            "role": "user",
                            "content": [
                                {"type": "image_url", "image_url": {
                                    "url": f"data:image/jpeg;base64,{b64}"}},
                                {"type": "text", "text": prompt},
                            ],
                        }],
                    },
                )
                data = resp.json()
                return data["choices"][0]["message"]["content"]
        except Exception as exc:
            logger.warning("GPT-4V failed: %s", exc)

    return "❌ 图片分析需要配置 ANTHROPIC_API_KEY 或 OPENAI_API_KEY"


async def _analyze_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from a file and query Aria LLM for analysis."""
    ext = Path(filename).suffix.lower()
    text_content = ""

    if ext == ".pdf":
        try:
            import pdfplumber, io
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                pages = [p.extract_text() or "" for p in pdf.pages[:20]]
                text_content = "\n".join(pages)[:8000]
        except ImportError:
            try:
                import pypdf, io
                reader = pypdf.PdfReader(io.BytesIO(file_bytes))
                text_content = "\n".join(
                    p.extract_text() or "" for p in reader.pages[:20]
                )[:8000]
            except Exception as exc:
                return f"❌ PDF 解析失败: {exc}"

    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            rows = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                for row in list(ws.iter_rows(values_only=True))[:50]:
                    rows.append("\t".join(str(c) for c in row if c is not None))
            text_content = f"[Excel: {filename}]\n" + "\n".join(rows)[:6000]
        except Exception as exc:
            return f"❌ Excel 解析失败: {exc}"

    elif ext in (".docx",):
        try:
            import docx, io
            doc = docx.Document(io.BytesIO(file_bytes))
            text_content = "\n".join(p.text for p in doc.paragraphs)[:8000]
        except Exception as exc:
            return f"❌ Word 解析失败: {exc}"

    elif ext in (".py", ".js", ".ts", ".go", ".java", ".cpp", ".c", ".rs",
                 ".txt", ".md", ".json", ".yaml", ".yml", ".csv", ".toml"):
        try:
            text_content = file_bytes.decode("utf-8", errors="replace")[:8000]
        except Exception:
            return "❌ 文件编码无法识别"

    else:
        try:
            text_content = file_bytes.decode("utf-8", errors="replace")[:4000]
        except Exception:
            return f"❌ 不支持的文件类型: {ext}"

    if not text_content.strip():
        return "❌ 文件内容为空或无法提取文本"

    query = (
        f"请分析以下文件（{filename}）的内容，给出主要信息、关键数据和洞察总结：\n\n"
        f"```\n{text_content[:6000]}\n```"
    )
    return await _query_aria_llm(query, timeout=120)


# ── Event verifier ────────────────────────────────────────────────────────────

def verify_feishu_signature(timestamp: str, nonce: str, body_bytes: bytes,
                            encrypt_key: str) -> bool:
    """Verify Feishu event signature (optional but recommended in production)."""
    if not encrypt_key:
        return True
    s = (timestamp + nonce + encrypt_key).encode() + body_bytes
    return hmac.compare_digest(
        hashlib.sha256(s).hexdigest(),
        ""  # caller should pass the X-Lark-Signature header value
    )


# ── Main event dispatcher (called by feishu_routes.py or standalone) ──────────

async def dispatch_event(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Handle one Feishu event payload.
    Supports: text / audio / image / file / post (富文本)
    Returns a dict to be sent as JSON response (HTTP 200 required by Feishu).
    """
    # 1. URL verification challenge
    if "challenge" in raw:
        return {"challenge": raw["challenge"]}

    header     = raw.get("header", {})
    event      = raw.get("event", {})
    event_type = header.get("event_type") or raw.get("type", "")

    if event_type not in ("im.message.receive_v1", "message"):
        return {"code": 0}

    msg      = event.get("message") or {}
    msg_id   = msg.get("message_id", "")
    msg_type = msg.get("message_type", "text")
    sender   = event.get("sender", {}).get("sender_id", {})
    user_id  = sender.get("user_id", "") or sender.get("open_id", "")
    chat_id  = msg.get("chat_id", "")

    logger.info("dispatch_event: type=%s msg_id=%s msg_type=%s user=%s chat=%s",
                event_type, msg_id, msg_type, user_id, chat_id)

    if not msg_id:
        logger.error("dispatch_event: msg_id is EMPTY — event structure may differ: %s",
                     json.dumps(raw, ensure_ascii=False)[:600])
        return {"code": 0}

    if not _is_allowed_user(user_id):
        logger.warning("Blocked user %s (not in FEISHU_ALLOWED_USER_IDS)", user_id)
        return {"code": 0}

    content_raw = msg.get("content", "{}")
    try:
        content = json.loads(content_raw)
    except Exception:
        content = {}

    # ── Text message ──────────────────────────────────────────────────────────
    if msg_type == "text":
        text = content.get("text", "").strip()
        # Strip @bot mention (飞书群里 @ 机器人会带前缀)
        if text.startswith("@"):
            text = " ".join(text.split()[1:]).strip()
        if not text:
            return {"code": 0}

        if text.startswith("/"):
            logger.info("Feishu /cmd from %s: %s", user_id, text[:80])
            asyncio.create_task(_handle_command(text, msg_id, user_id, chat_id))
        else:
            # Free-form natural language → Aria LLM
            logger.info("Feishu NL query from %s: %s", user_id, text[:80])
            asyncio.create_task(_handle_nl_query(text, msg_id, chat_id))

    # ── Voice / Audio ─────────────────────────────────────────────────────────
    elif msg_type == "audio":
        file_key = content.get("file_key", "")
        logger.info("Feishu audio from %s, key=%s", user_id, file_key)
        asyncio.create_task(_handle_audio(file_key, msg_id))

    # ── Image ────────────────────────────────────────────────────────────────
    elif msg_type == "image":
        image_key = content.get("image_key", "")
        logger.info("Feishu image from %s, key=%s", user_id, image_key)
        asyncio.create_task(_handle_image(image_key, msg_id))

    # ── File attachment ───────────────────────────────────────────────────────
    elif msg_type == "file":
        file_key  = content.get("file_key", "")
        file_name = content.get("file_name", "attachment")
        logger.info("Feishu file from %s: %s", user_id, file_name)
        asyncio.create_task(_handle_file(file_key, file_name, msg_id))

    # ── Rich text (post) — extract plain text ─────────────────────────────────
    elif msg_type == "post":
        try:
            # post content: {"zh_cn": {"title":"...","content":[[{"tag":"text","text":"..."},...]]}}
            lang_content = content.get("zh_cn") or content.get("en_us") or {}
            title = lang_content.get("title", "")
            paras = lang_content.get("content", [])
            texts = []
            for para in paras:
                for seg in para:
                    if seg.get("tag") == "text":
                        texts.append(seg.get("text", ""))
            plain = title + ("\n" if title else "") + " ".join(texts)
            if plain.strip():
                asyncio.create_task(_handle_nl_query(plain.strip(), msg_id))
        except Exception as exc:
            logger.warning("post parse error: %s", exc)

    return {"code": 0}


# ── Message type handlers (run as background tasks) ───────────────────────────

_MARKET_BRIEF_TRIGGERS = frozenset({
    "行情查询", "市场行情", "股市行情", "今日行情", "行情", "大盘", "大盘行情",
    "市场概况", "今日市场", "A股行情", "港股行情", "美股行情", "晨报",
    "market", "market overview", "stock market", "markets today",
})

_FOOTBALL_PREDICT_TRIGGERS = (
    "预测", "谁赢", "谁会赢", "比分预测", "胜率", "分析比赛",
    "predict", "who wins", "match preview",
)

# ── 常用 A股/港股 中文名 → 股票代码（用于 NL 解析）──────────────────────────
_CN_COMPANY_TICKER: dict[str, str] = {
    # 银行
    "工商银行": "601398", "工行": "601398",
    "建设银行": "601939", "建行": "601939",
    "农业银行": "601288", "农行": "601288",
    "中国银行": "601988", "中行": "601988",
    "招商银行": "600036", "招行": "600036",
    "平安银行": "000001",
    "兴业银行": "601166",
    "浦发银行": "600000",
    "光大银行": "601818",
    # 券商/保险
    "中信证券": "600030",
    "海通证券": "600837",
    "中国平安": "601318", "平安": "601318",
    "中国人寿": "601628",
    # 能源/化工
    "中国石油": "601857", "中石油": "601857",
    "中国石化": "600028", "中石化": "600028",
    "中国神华": "601088",
    # 有色金属
    "江西铜业": "600362",
    "紫金矿业": "601899",
    "中金黄金": "600489",
    "山东黄金": "600547",
    "洛阳钼业": "603993",
    "铜陵有色": "000630",
    "中国铝业": "601600",
    "南方铜业": "SCCO",
    "自由港": "FCX", "自由港麦克莫兰": "FCX",
    # 消费
    "贵州茅台": "600519", "茅台": "600519",
    "五粮液": "000858",
    "洋河股份": "002304",
    "海天味业": "603288",
    "伊利股份": "600887",
    "格力电器": "000651", "格力": "000651",
    "美的集团": "000333", "美的": "000333",
    "海尔智家": "600690",
    # 科技/互联网
    "腾讯": "0700", "腾讯控股": "0700",
    "阿里巴巴": "BABA", "阿里": "BABA",
    "京东": "JD",
    "百度": "BIDU",
    "比亚迪": "002594",
    "宁德时代": "300750", "宁德": "300750",
    "中芯国际": "688981",
    "海康威视": "002415",
    # 地产
    "万科": "000002", "万科A": "000002",
    "碧桂园": "2007",
    "恒大": "3333",
    # 医药
    "恒瑞医药": "600276",
    "迈瑞医疗": "300760",
    # 全球大宗商品矿企
    "必和必拓": "BHP", "必拓": "BHP",
    "力拓": "RIO", "力拓集团": "RIO",
    "淡水河谷": "VALE",
    "嘉能可": "GLEN.L",
}

# ── 大宗商品关键词 → (商品期货代码, 相关上市公司) ─────────────────────────────
_COMMODITY_MAP: dict[str, tuple[str, list[str], str]] = {
    # keyword: (yfinance_symbol, [related_tickers], display_name)
    "铜":    ("HG=F",  ["FCX", "SCCO", "600362", "601899", "000630"], "铜 COMEX"),
    "黄金":  ("GC=F",  ["GLD", "NEM", "GOLD", "600547", "600489"],   "黄金 COMEX"),
    "gold":  ("GC=F",  ["GLD", "NEM", "GOLD", "600547"],              "Gold COMEX"),
    "白银":  ("SI=F",  ["SLV", "PAAS", "AG"],                        "白银 COMEX"),
    "原油":  ("CL=F",  ["XOM", "CVX", "601857", "600028"],           "原油 WTI"),
    "oil":   ("CL=F",  ["XOM", "CVX", "BP", "601857"],               "Crude Oil WTI"),
    "天然气":("NG=F",  ["UNG", "LNG", "CQP"],                        "天然气 NYMEX"),
    "铁矿石":("TIO=F", ["BHP", "RIO", "VALE", "601088"],             "铁矿石"),
    "铝":    ("ALI=F", ["AA", "CENX", "601600"],                     "铝 LME"),
    "锂":    ("",      ["ALB", "SQM", "LTHM", "300750", "002594"],   "锂矿/电池"),
    "锂矿":  ("",      ["ALB", "SQM", "LTHM", "300750"],            "锂矿"),
    "小麦":  ("ZW=F",  ["ADM", "BG", "INGR"],                       "小麦 CBOT"),
    "大豆":  ("ZS=F",  ["ADM", "BG", "DE"],                         "大豆 CBOT"),
}


async def _fetch_market_snapshot() -> str:
    """
    Build a real-time market snapshot for A股 + 港股 + US indices directly via
    yfinance — bypasses the LLM subprocess so we always get actual data.
    """
    import yfinance as yf
    import asyncio

    _INDICES = [
        ("^SSEC",  "上证指数"),
        ("^HSI",   "恒生指数"),
        ("^GSPC",  "标普500"),
        ("^IXIC",  "纳斯达克"),
        ("^N225",  "日经225"),
        ("GC=F",   "黄金"),
        ("CL=F",   "原油"),
    ]

    def _fetch_one(sym: str):
        try:
            ti = yf.Ticker(sym)
            fi = ti.fast_info
            p  = getattr(fi, "last_price", None)
            pc = getattr(fi, "previous_close", None)
            if p and pc and pc > 0:
                pct = (p - pc) / pc * 100
                arrow = "▲" if pct >= 0 else "▼"
                return f"{arrow} {p:,.2f}  ({pct:+.2f}%)"
            elif p:
                return f"¥{p:,.2f}"
        except Exception:
            pass
        return "—"

    loop = asyncio.get_event_loop()
    lines = ["**主要市场行情**\n"]
    for sym, label in _INDICES:
        val = await loop.run_in_executor(None, _fetch_one, sym)
        lines.append(f"**{label}**　{val}")

    from datetime import datetime
    ts = datetime.now().strftime("%H:%M")
    lines.append(f"\n_更新时间: {ts}_")
    return "\n".join(lines)


async def _fetch_commodity_with_stocks(keyword: str) -> str:
    """
    Fetch commodity futures price + related stock prices for a given commodity keyword.
    Returns a formatted multi-line string for the Feishu card body.
    """
    import yfinance as yf
    import asyncio

    entry = _COMMODITY_MAP.get(keyword)
    if not entry:
        return ""
    futures_sym, related_tickers, display_name = entry

    def _price_line(sym: str) -> str:
        try:
            ti = yf.Ticker(sym)
            fi = ti.fast_info
            p  = getattr(fi, "last_price", None)
            pc = getattr(fi, "previous_close", None)
            name = sym
            # Try to get a short name
            try:
                info = ti.info
                name = info.get("shortName", sym)[:12]
            except Exception:
                pass
            if p and pc and pc > 0:
                pct = (p - pc) / pc * 100
                arrow = "▲" if pct >= 0 else "▼"
                currency = "¥" if sym.isdigit() else "$"
                return f"{arrow} **{name}** ({sym})  {currency}{p:,.2f}  ({pct:+.2f}%)"
            elif p:
                currency = "¥" if sym.isdigit() else "$"
                return f"**{name}** ({sym})  {currency}{p:,.2f}"
        except Exception:
            pass
        return f"**{sym}**  —"

    loop = asyncio.get_event_loop()
    sections = []

    # Commodity futures
    if futures_sym:
        fut_line = await loop.run_in_executor(None, _price_line, futures_sym)
        sections.append(f"**{display_name} 期货**\n{fut_line}")

    # Related stocks
    stock_lines = []
    for sym in related_tickers[:5]:
        # A-share: append exchange suffix for yfinance
        yfn = sym
        if sym.isdigit() and len(sym) == 6:
            yfn = sym + (".SS" if sym.startswith(("6", "5")) else ".SZ")
        line = await loop.run_in_executor(None, _price_line, yfn)
        stock_lines.append(line)

    if stock_lines:
        sections.append("**相关上市公司**\n" + "\n".join(stock_lines))

    from datetime import datetime
    ts = datetime.now().strftime("%H:%M")
    return "\n\n".join(sections) + f"\n\n_数据时间: {ts}_"


def _resolve_cn_company(text: str) -> str:
    """
    Replace Chinese company names in text with their ticker symbols.
    E.g. "江西铜业的走势" → "600362的走势"
    Returns modified text; if nothing matched, returns original.
    """
    result = text
    for cn_name, ticker in _CN_COMPANY_TICKER.items():
        if cn_name in result:
            result = result.replace(cn_name, ticker)
    return result


async def _handle_nl_query(text: str, message_id: str, chat_id: str = "") -> None:
    """Route free-form natural language to Aria LLM and reply."""
    import re as _re_nl
    _low = text.strip().lower()
    _orig = text.strip()

    # Fast-path: generic market brief request — call yfinance directly (bypasses LLM)
    if _low in _MARKET_BRIEF_TRIGGERS or (
        any(k in _low for k in ("行情查询", "市场行情", "大盘今日", "今日大盘")) and
        not any(c.isalpha() and c.upper() == c for c in text)  # no uppercase ticker
    ):
        await reply_or_send(message_id, chat_id, "📊 获取市场概况…", "正在抓取主要指数数据…", "blue")
        try:
            snapshot = await _fetch_market_snapshot()
            await reply_or_send(message_id, chat_id, "📊 市场行情", snapshot, "turquoise",
                                footer="Aria Code · yfinance 实时数据")
        except Exception as _e:
            await reply_or_send(message_id, chat_id, "❌ 行情获取失败", str(_e)[:200], "red")
        return

    # Fast-path: commodity + related stocks (e.g. "铜的相关公司", "黄金走势")
    _commodity_kw = None
    for _ck in _COMMODITY_MAP:
        if _ck in _low:
            _commodity_kw = _ck
            break
    _needs_stocks = any(k in _low for k in ("相关公司", "相关股票", "产业链", "概念股", "走势", "行情", "估值"))
    if _commodity_kw and (_needs_stocks or len(_orig) <= 6):
        await reply_or_send(message_id, chat_id, f"🔍 查询{_commodity_kw}行情…",
                            "正在获取期货及相关股票数据…", "blue")
        body = await _fetch_commodity_with_stocks(_commodity_kw)
        if body:
            await reply_or_send(message_id, chat_id, f"📦 {_commodity_kw} 市场概况",
                                body, "turquoise", footer="Aria Code · 大宗商品数据")
            return

    # Chinese company name → ticker resolution before sending to LLM
    resolved = _resolve_cn_company(_orig)
    if resolved != _orig:
        # Found at least one CN name; also annotate so LLM has context
        text = resolved + f"  (原文: {_orig})"

    # Fast-path: football match prediction — extract team names and call Poisson model directly
    _has_predict_kw = any(k in _low for k in _FOOTBALL_PREDICT_TRIGGERS)
    _has_football_kw = any(k in _low for k in ("足球", "世界杯", "欧冠", "英超", "比赛", "football", "soccer", "match", "world cup"))
    if _has_predict_kw or _has_football_kw:
        # Strip one or more leading context words (handles "预测今天加拿大跟波黑")
        _stripped = _re_nl.sub(
            r'^(?:预测|分析|谁赢|谁会赢|今天|明天|比赛|足球|世界杯|结果|比分|\s)+',
            '', _orig, flags=_re_nl.IGNORECASE
        )
        _vs_m = _re_nl.search(
            r'(.{2,20}?)\s*(?:vs\.?\s*|对阵\s*|对\s+|跟\s*|和\s*|pk\s*|——\s*|—\s*)(.{2,20})',
            _stripped, _re_nl.IGNORECASE
        )
        if _vs_m:
            home_t = _vs_m.group(1).strip().rstrip("的在")
            away_t = _vs_m.group(2).strip().rstrip("的在")
            if home_t and away_t and len(home_t) >= 2 and len(away_t) >= 2:
                await reply_or_send(message_id, chat_id, "⚽ 预测中…",
                                    f"> {home_t} vs {away_t}", "blue")
                await _handle_football_predict(f"{home_t} vs {away_t}", message_id)
                return

    await reply_or_send(message_id, chat_id, "🤔 思考中…", f"> {_orig[:120]}", "blue")
    # Use direct LLM call (no subprocess, no tool execution) for conversational queries
    result = await _query_aria_direct(text, timeout=120)
    color = "red" if result.startswith("❌") else "green"
    await reply_or_send(message_id, chat_id, "💡 Aria 回答", result[:2000], color,
                        footer="Aria Code · AI 分析")


async def _handle_audio(file_key: str, message_id: str) -> None:
    """Download voice → transcribe → query Aria LLM."""
    await reply_card(message_id, "🎤 正在转写语音…", "下载中，请稍候…", "blue")
    audio_bytes = await _download_feishu_resource(message_id, file_key, rtype="file")
    if not audio_bytes:
        await reply_card(message_id, "❌ 语音下载失败", "无法获取语音文件", "red")
        return
    text = await _transcribe_voice(audio_bytes)
    if text.startswith("❌"):
        await reply_card(message_id, "❌ 语音转文字失败", text, "red")
        return
    await reply_card(message_id, "🎤 识别结果", f"**语音内容：**\n{text}\n\n---\n正在分析…", "blue")
    result = await _query_aria_llm(text, timeout=120)
    await reply_card(message_id, "💡 Aria 回答", result[:2000], "green",
                     footer=f"语音转文字: {text[:60]}…")


async def _handle_image(image_key: str, message_id: str) -> None:
    """Download image → visual LLM analysis."""
    await reply_card(message_id, "🖼️ 分析图片中…", "正在下载并识别，请稍候…", "blue")
    img_bytes = await _download_feishu_resource(message_id, image_key, rtype="image")
    if not img_bytes:
        await reply_card(message_id, "❌ 图片下载失败", "无法获取图片", "red")
        return
    result = await _analyze_image(img_bytes)
    await reply_card(message_id, "🖼️ 图片分析", result[:2000], "turquoise",
                     footer="Aria Code · 视觉 AI")


async def _handle_file(file_key: str, filename: str, message_id: str) -> None:
    """Download file → parse → Aria LLM analysis."""
    await reply_card(message_id, f"📄 解析文件：{filename}",
                     "正在下载并解析，请稍候…", "blue")
    file_bytes = await _download_feishu_resource(message_id, file_key, rtype="file")
    if not file_bytes:
        await reply_card(message_id, "❌ 文件下载失败", "无法获取文件", "red")
        return
    result = await _analyze_file(file_bytes, filename)
    await reply_card(message_id, f"📄 {filename} 分析完成", result[:2000], "turquoise",
                     footer="Aria Code · 文件智能解析")


# ── Standalone HTTP server (for testing without FastAPI) ──────────────────────

async def _standalone_server(host: str = "0.0.0.0", port: int = 8888) -> None:
    """Minimal aiohttp-based server for standalone Feishu event reception."""
    try:
        from aiohttp import web
    except ImportError:
        logger.error("aiohttp not installed. pip install aiohttp")
        return

    async def handle(request):
        try:
            body = await request.json()
        except Exception:
            return web.json_response({"code": 1, "msg": "bad json"}, status=400)
        result = await dispatch_event(body)
        return web.json_response(result)

    app = web.Application()
    app.router.add_post("/feishu/event", handle)
    app.router.add_post("/api/v1/feishu/event", handle)

    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, host, port)
    await site.start()
    logger.info("Feishu standalone server listening on http://%s:%d/feishu/event", host, port)
    logger.info("Configure this URL in Feishu Developer Console → Event Subscription")
    await asyncio.Event().wait()


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [feishu] %(levelname)s %(message)s")
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8888
    asyncio.run(_standalone_server(port=port))
