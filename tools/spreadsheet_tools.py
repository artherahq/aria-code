"""
spreadsheet_tools.py — 结构化 Excel 工作簿生成 + 独立校验
==========================================================
输入: 结构化 spec（多工作表 / 表头 / 数据行 / 合计声明）
输出: 带样式与公式的 .xlsx（默认落盘到 artifacts 产物目录）

设计原则（与 /report 研报引擎同一哲学：确定性优先）:
  · **公式由工具生成，不由模型手写** —— 模型只声明意图
    （"对 F 列按 E 列=GBP 条件求和"），SUMIFS/COUNTIF 的单元格范围
    由工具按实际数据行数计算，从根源上消灭"范围错位"这类静默错误。
  · **Python 独立复核** —— 每个合计公式落盘的同时，用纯 Python 对
    同一数据重算一遍，结果随返回值给出；两条路径殊途同归才算可信。
  · **禁止硬编码合计** —— 合计单元格只写公式，Excel/Numbers 打开时
    即时计算，数据行修改后合计自动跟随。
  · 样式约定内置（Arial、深色表头、千分位、冻结首行、标黄高亮），
    模型无需（也不能）逐单元格调样式。

对外暴露:
  write_workbook(spec, out_path=None)   → dict（路径 + 校验结果）
  tool_write_spreadsheet(params)        → Agent 工具入口（JSON 安全）
  register_spreadsheet_tools(tools, schemas) → 注册进 LOCAL_TOOLS

完全离线，无网络依赖；openpyxl 已在项目基础依赖中。
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - openpyxl 在 pyproject 基础依赖中
    HAS_OPENPYXL = False

# ── 样式约定（全局统一，勿逐表覆盖） ─────────────────────────────────────────
_FONT = "Arial"
_HDR_FILL = "1F2937"       # 深灰蓝表头
_HIGHLIGHT_FILL = "FEF3C7" # 标黄：假设 / 待确认条目
_BORDER_COLOR = "D1D5DB"
_DATE_FMT = "yyyy-mm-dd"
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# Agent 工具防线（防止一次调用写出百万行拖死 REPL）
MAX_SHEETS = 12
MAX_ROWS_PER_SHEET = 5000
MAX_COLS = 40

_TOTAL_TYPES = ("sum", "sumif", "count", "countif", "average")


def _styles():
    thin = Side(style="thin", color=_BORDER_COLOR)
    return {
        "hdr_font": Font(name=_FONT, bold=True, color="FFFFFF", size=11),
        "hdr_fill": PatternFill("solid", fgColor=_HDR_FILL),
        "body": Font(name=_FONT, size=10),
        "bold": Font(name=_FONT, bold=True, size=10),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "hl_fill": PatternFill("solid", fgColor=_HIGHLIGHT_FILL),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(vertical="center", wrap_text=True),
    }


def _coerce_cell(value: Any) -> Any:
    """ISO 日期字符串 → date；其余原样。"""
    if isinstance(value, str) and _ISO_DATE.match(value):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    return value


def _col_idx(letter: str) -> int:
    return column_index_from_string(letter.upper())


# ── 合计公式：生成 + Python 复核（同一份声明，两条独立路径） ────────────────

def _total_formula(total: Dict[str, Any], first_row: int, last_row: int) -> str:
    kind = str(total.get("type", "sum")).lower()
    col = str(total["value_col"]).upper()
    rng = f"{col}{first_row}:{col}{last_row}"
    if kind == "sum":
        return f"=SUM({rng})"
    if kind == "average":
        return f"=AVERAGE({rng})"
    if kind == "count":
        return f"=COUNTA({rng})"
    key_col = str(total["key_col"]).upper()
    key = str(total["key"])
    key_rng = f"{key_col}{first_row}:{key_col}{last_row}"
    if kind == "sumif":
        return f'=SUMIFS({rng},{key_rng},"{key}")'
    if kind == "countif":
        return f'=COUNTIF({key_rng},"{key}")'
    raise ValueError(f"unsupported total type: {kind}")


def _total_python(total: Dict[str, Any], rows: List[List[Any]]) -> float:
    """对同一声明用纯 Python 重算 —— 与公式路径互为独立校验。"""
    kind = str(total.get("type", "sum")).lower()
    vi = _col_idx(str(total["value_col"])) - 1

    def _nums(filtered):
        for r in filtered:
            v = r[vi] if vi < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                yield v

    if kind in ("sum", "average", "count"):
        picked = rows
    else:
        ki = _col_idx(str(total["key_col"])) - 1
        key = str(total["key"])
        picked = [r for r in rows if ki < len(r) and str(r[ki]) == key]

    if kind == "count":
        return float(sum(1 for r in picked if vi < len(r) and r[vi] not in (None, "")))
    if kind == "countif":
        return float(len(picked))
    vals = list(_nums(picked))
    if kind == "average":
        return round(sum(vals) / len(vals), 6) if vals else 0.0
    return round(sum(vals), 6)


# ── 核心：spec → Workbook ────────────────────────────────────────────────────

def build_workbook(spec: Dict[str, Any]) -> Tuple["Workbook", List[Dict[str, Any]]]:
    """按 spec 构建 Workbook。返回 (wb, 校验清单)。

    sheet spec 字段:
      name            工作表名（必填）
      headers         表头列表（必填）
      rows            数据行（二维数组；ISO 日期字符串自动转日期）
      number_formats  {"F": "#,##0.00"} 列号→格式（可选）
      col_widths      [宽度…] 与表头对齐（可选）
      freeze_header   冻结首行，默认 True
      highlight_rows  [1,5] 数据区内 1-based 行号，标黄（可选）
      totals          合计声明列表（可选），元素:
                      {label, type: sum|sumif|count|countif|average,
                       value_col: "F", key_col: "E", key: "GBP",
                       label_col: "D"}
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl 未安装（应随基础依赖提供）")

    st = _styles()
    wb = Workbook()
    wb.remove(wb.active)
    verifications: List[Dict[str, Any]] = []

    for sheet in spec.get("sheets", []):
        name = str(sheet["name"])[:31]  # Excel 工作表名 31 字符上限
        ws = wb.create_sheet(name)
        headers: List[str] = list(sheet.get("headers", []))
        rows: List[List[Any]] = [list(r) for r in sheet.get("rows", [])]

        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = st["hdr_font"], st["hdr_fill"]
            c.alignment, c.border = st["center"], st["border"]

        highlight = {int(i) for i in sheet.get("highlight_rows", [])}
        for i, row in enumerate(rows, start=1):
            ws.append([_coerce_cell(v) for v in row])
            r = ws.max_row
            for c in ws[r]:
                c.font, c.border, c.alignment = st["body"], st["border"], st["wrap"]
            if i in highlight:
                for c in ws[r]:
                    c.fill = st["hl_fill"]

        fmts: Dict[str, str] = dict(sheet.get("number_formats", {}))
        for letter, fmt in fmts.items():
            ci = _col_idx(letter)
            for r in range(2, len(rows) + 2):
                ws.cell(r, ci).number_format = fmt
        # 日期列自动格式（凡 coerce 出 date 的单元格）
        for r in range(2, len(rows) + 2):
            for c in ws[r]:
                if isinstance(c.value, (date, datetime)) and not c.number_format.startswith("yy"):
                    c.number_format = _DATE_FMT

        first_row, last_row = 2, len(rows) + 1
        totals = list(sheet.get("totals", []))
        if totals and rows:
            ws.append([])
            for total in totals:
                blank = [""] * len(headers)
                label_ci = _col_idx(str(total.get("label_col", "A"))) - 1
                value_ci = _col_idx(str(total["value_col"])) - 1
                blank[label_ci] = str(total.get("label", "合计"))
                ws.append(blank)
                r = ws.max_row
                formula = _total_formula(total, first_row, last_row)
                ws.cell(r, value_ci + 1, formula)
                for c in ws[r]:
                    c.font = st["bold"]
                fmt = fmts.get(str(total["value_col"]).upper())
                if fmt:
                    ws.cell(r, value_ci + 1).number_format = fmt
                verifications.append({
                    "sheet": name,
                    "label": str(total.get("label", "合计")),
                    "formula": formula,
                    "python_value": _total_python(total, [
                        [_coerce_cell(v) for v in row] for row in rows
                    ]),
                })

        widths = sheet.get("col_widths") or []
        for i, w in enumerate(widths[:len(headers)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = int(w)
        if sheet.get("freeze_header", True):
            ws.freeze_panes = "A2"

    return wb, verifications


def write_workbook(spec: Dict[str, Any], out_path: Optional[Path] = None) -> Dict[str, Any]:
    """构建并落盘。out_path 缺省时写入用户产物目录（artifacts）。"""
    wb, verifications = build_workbook(spec)

    if out_path is None:
        stem = str(spec.get("filename") or "workbook").removesuffix(".xlsx")
        try:
            from artifacts import create_user_artifact, write_artifact_metadata
            record = create_user_artifact(
                "spreadsheets", spec.get("topic"), stem, ".xlsx"
            )
            out_path = record.path
            wb.save(out_path)
            write_artifact_metadata(record, {
                "generator": "spreadsheet_tools.write_workbook",
                "sheets": [s.get("name") for s in spec.get("sheets", [])],
                "verified_totals": verifications,
            })
        except Exception as exc:  # artifacts 不可用时退化到当前目录
            logger.debug("[xlsx] artifacts unavailable, fallback cwd: %s", exc)
            out_path = Path.cwd() / f"{stem}.xlsx"
            wb.save(out_path)
    else:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

    return {
        "success": True,
        "path": str(out_path),
        "sheets": [
            {"name": s.get("name"), "rows": len(s.get("rows", []))}
            for s in spec.get("sheets", [])
        ],
        "verified_totals": verifications,
        "note": "合计单元格为公式，Excel/Numbers 打开时即时计算；"
                "python_value 为独立复核结果，应与打开后的公式值一致。",
    }


# ── Agent 工具入口 ───────────────────────────────────────────────────────────

def tool_write_spreadsheet(params: Dict[str, Any]) -> Dict[str, Any]:
    """LOCAL_TOOLS 入口：校验尺寸防线 → write_workbook。永不抛异常。"""
    try:
        sheets = params.get("sheets") or []
        if not sheets:
            return {"success": False, "error": "spec 缺少 sheets"}
        if len(sheets) > MAX_SHEETS:
            return {"success": False, "error": f"工作表数超上限 {MAX_SHEETS}"}
        for s in sheets:
            if len(s.get("rows", [])) > MAX_ROWS_PER_SHEET:
                return {"success": False,
                        "error": f"'{s.get('name')}' 行数超上限 {MAX_ROWS_PER_SHEET}"}
            if len(s.get("headers", [])) > MAX_COLS:
                return {"success": False,
                        "error": f"'{s.get('name')}' 列数超上限 {MAX_COLS}"}
            for t in s.get("totals", []):
                if str(t.get("type", "sum")).lower() not in _TOTAL_TYPES:
                    return {"success": False,
                            "error": f"不支持的合计类型: {t.get('type')}"}
        return write_workbook(params)
    except Exception as exc:
        logger.debug("[xlsx] tool_write_spreadsheet error: %s", exc, exc_info=True)
        return {"success": False, "error": str(exc)}


SPREADSHEET_TOOL_SCHEMAS: List[Dict[str, Any]] = [
    {
        "type": "function",
        "function": {
            "name": "write_spreadsheet",
            "description": (
                "生成带样式与公式的多工作表 Excel(.xlsx)，适合对账单/费用清单/"
                "持仓汇总等结构化交付物。合计(SUM/SUMIFS/COUNTIF/AVERAGE)只需声明"
                "意图，单元格范围由工具按数据行数自动生成并用 Python 独立复核，"
                "返回值中的 verified_totals 给出复核结果。日期请用 ISO 字符串"
                "(YYYY-MM-DD)。文件写入用户产物目录并返回路径。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "文件名(不含扩展名)"},
                    "topic": {"type": "string", "description": "产物子目录主题(可选)"},
                    "sheets": {
                        "type": "array",
                        "description": "工作表列表(按顺序)",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "headers": {"type": "array", "items": {"type": "string"}},
                                "rows": {
                                    "type": "array",
                                    "items": {"type": "array"},
                                    "description": "二维数据；数字用 number，日期用 YYYY-MM-DD 字符串",
                                },
                                "number_formats": {
                                    "type": "object",
                                    "description": '列格式，如 {"F": "#,##0.00"}',
                                },
                                "col_widths": {
                                    "type": "array", "items": {"type": "integer"},
                                    "description": "各列宽度，与表头对齐",
                                },
                                "highlight_rows": {
                                    "type": "array", "items": {"type": "integer"},
                                    "description": "需标黄的数据行号(1-based)，用于假设/待确认条目",
                                },
                                "totals": {
                                    "type": "array",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "label": {"type": "string"},
                                            "type": {"type": "string",
                                                     "enum": list(_TOTAL_TYPES)},
                                            "value_col": {"type": "string",
                                                          "description": "求值列字母，如 F"},
                                            "key_col": {"type": "string",
                                                        "description": "sumif/countif 的条件列字母"},
                                            "key": {"type": "string",
                                                    "description": "sumif/countif 的匹配值"},
                                            "label_col": {"type": "string",
                                                          "description": "合计标签落在哪一列，默认 A"},
                                        },
                                        "required": ["type", "value_col"],
                                    },
                                },
                            },
                            "required": ["name", "headers", "rows"],
                        },
                    },
                },
                "required": ["sheets"],
            },
        },
    },
]


def register_spreadsheet_tools(tool_registry: Dict, schema_registry: List) -> int:
    """注册进 LOCAL_TOOLS / LOCAL_TOOL_SCHEMAS（从不覆盖已有工具）。"""
    if not HAS_OPENPYXL:
        return 0
    added = 0
    if "write_spreadsheet" not in tool_registry:
        tool_registry["write_spreadsheet"] = (
            tool_write_spreadsheet,
            "生成带公式与样式的多工作表 Excel 文件",
        )
        added += 1
    existing = {s.get("function", {}).get("name") for s in schema_registry}
    for schema in SPREADSHEET_TOOL_SCHEMAS:
        if schema["function"]["name"] not in existing:
            schema_registry.append(schema)
    return added
