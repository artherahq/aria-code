"""
data_analysis_tools.py — 数据分析增强层
==========================================
提供：
  - Excel 导出 (多 Sheet + 内嵌图表)
  - DuckDB SQL 接口（本地内存数据仓库）
  - 价格预警系统 (~/.arthera/alerts.json)
  - 多资产组合回测 (相关性矩阵 + 权重收益)
  - 相关性热力图数据
  - 自定义因子表达式 DSL

依赖（可选）：
    pip install openpyxl duckdb pandas numpy yfinance
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ── Optional imports ──────────────────────────────────────────────────────────

try:
    import pandas as pd
    _HAS_PD = True
except ImportError:
    _HAS_PD = False

try:
    import numpy as np
    _HAS_NP = True
except ImportError:
    _HAS_NP = False

try:
    import yfinance as yf
    _HAS_YF = True
except ImportError:
    _HAS_YF = False

try:
    import duckdb
    _HAS_DUCK = True
except ImportError:
    _HAS_DUCK = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

ALERTS_PATH = Path.home() / ".arthera" / "alerts.json"
EXPORT_DIR  = Path.home() / ".arthera" / "exports"


# ── Excel 导出 ────────────────────────────────────────────────────────────────

def export_to_excel(params: dict) -> dict:
    """
    将数据导出到格式化 Excel 文件（多 Sheet）。

    参数：
      data:      dict，键=Sheet名，值=list[dict] 行数据
      filename:  输出文件名（默认 aria_export_<timestamp>.xlsx）
      add_chart: 是否为含 '价格'/'close'/'收盘' 列的 Sheet 添加折线图
    """
    if not _HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl 未安装，请运行: pip install openpyxl"}
    if not _HAS_PD:
        return {"success": False, "error": "pandas 未安装"}

    data = params.get("data", {})
    if not data:
        return {"success": False, "error": "data 参数不能为空"}

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = params.get("filename") or f"aria_export_{ts}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    out_path = EXPORT_DIR / filename
    add_chart = bool(params.get("add_chart", True))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Style constants
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    alt_fill     = PatternFill("solid", fgColor="EBF5FB")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        bottom=Side(style="thin", color="BDC3C7"),
    )

    for sheet_name, rows in data.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=str(sheet_name)[:31])  # Excel max 31 chars
        df = pd.DataFrame(rows)
        cols = list(df.columns)

        # Header row
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=str(col))
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center_align

        # Data rows
        price_col_idx = None
        for ri, row_dict in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                val = row_dict.get(col)
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri % 2 == 0:
                    cell.fill = alt_fill
                cell.border = thin_border
                # Auto-detect price column
                col_lower = col.lower()
                if price_col_idx is None and any(k in col_lower for k in
                        ["价格","close","收盘","price","last","最新"]):
                    price_col_idx = ci

        # Auto-column width
        for ci, col in enumerate(cols, 1):
            max_len = max(len(str(col)),
                          *(len(str(r.get(col,""))) for r in rows[:50]))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 40)

        # Freeze header
        ws.freeze_panes = "A2"

        # Add line chart if price column detected
        if add_chart and price_col_idx and len(rows) >= 3:
            chart = LineChart()
            chart.title  = f"{sheet_name} 价格走势"
            chart.style  = 10
            chart.y_axis.title = "价格"
            chart.x_axis.title = "时间"
            data_ref = Reference(ws, min_col=price_col_idx,
                                  min_row=1, max_row=len(rows) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.width  = 20
            chart.height = 12
            ws.add_chart(chart, f"A{len(rows) + 4}")

    # Summary sheet
    ws_sum = wb.create_sheet(title="📊 汇总", index=0)
    ws_sum["A1"] = "Aria Code 数据导出报告"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_sum["A3"] = "导出时间："
    ws_sum["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_sum["A4"] = "Sheet 数量："
    ws_sum["B4"] = len(data)
    ws_sum["A5"] = "总行数："
    ws_sum["B5"] = sum(len(v) for v in data.values())
    ws_sum["A7"] = "包含 Sheet："
    for i, sname in enumerate(data.keys(), 8):
        ws_sum[f"A{i}"] = f"  • {sname}"
        ws_sum[f"B{i}"] = f"{len(data[sname])} 行"

    wb.save(str(out_path))
    return {
        "success":   True,
        "path":      str(out_path),
        "sheets":    list(data.keys()),
        "total_rows": sum(len(v) for v in data.values()),
        "filename":  filename,
    }


# ── DuckDB SQL 接口 ───────────────────────────────────────────────────────────

# Persistent in-memory DuckDB connection for the session
_duck_conn: Optional[Any] = None
_duck_tables: Dict[str, bool] = {}


def _get_duck_conn():
    global _duck_conn
    if _duck_conn is None:
        _duck_conn = duckdb.connect(":memory:")
    return _duck_conn


def sql_query(params: dict) -> dict:
    """
    在内存 DuckDB 中执行 SQL 查询。

    参数：
      query:   SQL 语句
      load:    可选，dict {table_name: list[dict]} — 在查询前加载数据
      limit:   结果行数限制（默认 500）
    """
    if not _HAS_DUCK:
        return {"success": False,
                "error": "duckdb 未安装，请运行: pip install duckdb"}
    if not _HAS_PD:
        return {"success": False, "error": "pandas 未安装"}

    query = str(params.get("query", "")).strip()
    if not query:
        return {"success": False, "error": "query 不能为空"}

    conn  = _get_duck_conn()
    limit = int(params.get("limit", 500))
    load  = params.get("load", {})

    # Load tables
    for tname, rows in (load or {}).items():
        if rows:
            df = pd.DataFrame(rows)
            conn.register(tname, df)
            _duck_tables[tname] = True

    # Safety check: block destructive operations
    q_upper = query.upper().lstrip()
    if any(q_upper.startswith(kw) for kw in ("DROP ", "DELETE ", "TRUNCATE ")):
        return {"success": False,
                "error": "安全限制：不允许 DROP/DELETE/TRUNCATE 操作"}

    # Auto-add LIMIT if SELECT without one
    if q_upper.startswith("SELECT") and "LIMIT" not in q_upper:
        query = f"{query.rstrip(';')} LIMIT {limit}"

    try:
        result = conn.execute(query).fetchdf()
        records = result.to_dict("records")
        cols    = list(result.columns)
        return {
            "success":  True,
            "columns":  cols,
            "rows":     records,
            "row_count": len(records),
            "tables_loaded": list(_duck_tables.keys()),
        }
    except Exception as e:
        return {"success": False, "error": str(e), "query": query}


def sql_list_tables(params: dict = None) -> dict:
    """列出内存 DuckDB 中已加载的表。"""
    if not _HAS_DUCK:
        return {"success": False, "error": "duckdb 未安装"}
    conn = _get_duck_conn()
    try:
        df = conn.execute("SHOW TABLES").fetchdf()
        tables = df["name"].tolist() if "name" in df.columns else []
        return {"success": True, "tables": tables}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── 价格预警系统 ──────────────────────────────────────────────────────────────

def _load_alerts() -> List[dict]:
    if ALERTS_PATH.exists():
        try:
            return json.loads(ALERTS_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []


def _save_alerts(alerts: list) -> None:
    ALERTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    ALERTS_PATH.write_text(json.dumps(alerts, ensure_ascii=False, indent=2),
                            encoding="utf-8")


def add_price_alert(params: dict) -> dict:
    """
    添加价格预警。

    参数：
      symbol:     标的代码（如 "AAPL" / "600519"）
      condition:  触发条件 ("gt" / "lt" / "cross_up" / "cross_down")
      price:      触发价格
      note:       备注（可选）
    """
    symbol    = str(params.get("symbol", "")).upper().strip()
    condition = str(params.get("condition", "gt")).lower()
    price     = float(params.get("price", 0))
    note      = str(params.get("note", ""))

    if not symbol:
        return {"success": False, "error": "symbol 不能为空"}
    if condition not in ("gt", "lt", "cross_up", "cross_down"):
        return {"success": False,
                "error": "condition 必须是 gt/lt/cross_up/cross_down"}
    if price <= 0:
        return {"success": False, "error": "price 必须 > 0"}

    alerts = _load_alerts()
    alert_id = f"{symbol}_{condition}_{price}_{int(datetime.now().timestamp())}"
    new_alert = {
        "id":        alert_id,
        "symbol":    symbol,
        "condition": condition,
        "price":     price,
        "note":      note,
        "created_at": datetime.now().isoformat(),
        "triggered": False,
    }
    alerts.append(new_alert)
    _save_alerts(alerts)

    cond_label = {"gt": ">", "lt": "<",
                  "cross_up": "向上突破", "cross_down": "向下跌破"}[condition]
    return {
        "success":  True,
        "alert_id": alert_id,
        "message":  f"已设置预警：{symbol} 价格 {cond_label} {price}",
        "total_alerts": len(alerts),
    }


def list_price_alerts(params: dict = None) -> dict:
    """列出所有价格预警。"""
    alerts = _load_alerts()
    active  = [a for a in alerts if not a.get("triggered")]
    done    = [a for a in alerts if a.get("triggered")]
    return {
        "success":        True,
        "active_alerts":  active,
        "triggered_alerts": done,
        "total":          len(alerts),
    }


def delete_price_alert(params: dict) -> dict:
    """删除指定 ID 的价格预警。"""
    alert_id = str(params.get("alert_id", ""))
    if not alert_id:
        return {"success": False, "error": "alert_id 不能为空"}
    alerts = _load_alerts()
    before = len(alerts)
    alerts = [a for a in alerts if a.get("id") != alert_id]
    if len(alerts) == before:
        return {"success": False, "error": f"未找到预警 {alert_id}"}
    _save_alerts(alerts)
    return {"success": True, "deleted_id": alert_id, "remaining": len(alerts)}


def check_alerts(params: dict = None) -> dict:
    """
    检查当前所有未触发预警的状态（需联网获取实时价格）。
    返回已触发的预警列表。
    """
    if not _HAS_YF:
        return {"success": False, "error": "yfinance 未安装，无法获取实时价格"}

    alerts  = _load_alerts()
    active  = [a for a in alerts if not a.get("triggered")]
    if not active:
        return {"success": True, "triggered": [], "message": "无活跃预警"}

    # Batch price fetch
    symbols = list(set(a.get("symbol", "") for a in active if a.get("symbol")))
    prices: Dict[str, float] = {}
    try:
        tickers = yf.Tickers(" ".join(symbols))
        for sym in symbols:
            try:
                info = tickers.tickers[sym].fast_info
                px = getattr(info, "last_price", None) or \
                     getattr(info, "regularMarketPrice", None)
                if px:
                    prices[sym] = float(px)
            except Exception:
                pass
    except Exception as e:
        logger.debug("Alert price fetch failed: %s", e)

    triggered_now = []
    for alert in active:
        sym  = alert.get("symbol")
        cond = alert.get("condition")
        tgt  = alert.get("price")
        if not sym or not cond or tgt is None:
            continue
        cur  = prices.get(sym)
        if cur is None:
            continue
        hit = False
        if cond == "gt" and cur > tgt:
            hit = True
        elif cond == "lt" and cur < tgt:
            hit = True
        elif cond == "cross_up" and cur >= tgt:
            hit = True
        elif cond == "cross_down" and cur <= tgt:
            hit = True
        if hit:
            alert["triggered"]    = True
            alert["triggered_at"] = datetime.now().isoformat()
            alert["triggered_price"] = cur
            triggered_now.append(alert)

    _save_alerts(alerts)

    # Push notifications for newly triggered alerts
    if triggered_now:
        try:
            from notification_tools import send_alert_notification
            for _alrt in triggered_now:
                send_alert_notification(_alrt)
        except Exception as _ne:
            logger.debug("Notification dispatch failed: %s", _ne)

    return {
        "success":   True,
        "triggered": triggered_now,
        "checked":   len(active),
        "prices":    prices,
    }


# ── 相关性矩阵 ────────────────────────────────────────────────────────────────

def calc_correlation_matrix(params: dict) -> dict:
    """
    计算多资产收益率相关性矩阵。

    参数：
      symbols:   list[str]，如 ["AAPL","MSFT","TSLA","SPY"]
      period:    历史区间 ("1y" / "2y" / "6mo"，默认 "1y")
      interval:  频率 ("1d" / "1wk"，默认 "1d")
    """
    if not _HAS_YF:
        return {"success": False, "error": "yfinance 未安装"}
    if not _HAS_PD or not _HAS_NP:
        return {"success": False, "error": "pandas / numpy 未安装"}

    symbols  = params.get("symbols", [])
    period   = str(params.get("period", "1y"))
    interval = str(params.get("interval", "1d"))

    if len(symbols) < 2:
        return {"success": False, "error": "至少需要 2 个标的"}
    if len(symbols) > 20:
        symbols = symbols[:20]

    try:
        raw = yf.download(symbols, period=period, interval=interval,
                          progress=False, auto_adjust=True)
        if raw.empty:
            return {"success": False, "error": "下载数据为空"}

        if isinstance(raw.columns, pd.MultiIndex):
            closes = raw["Close"]
        else:
            closes = raw

        closes = closes.dropna(how="all", axis=1)
        rets   = closes.pct_change().dropna()
        corr   = rets.corr()

        # Convert to serializable format
        corr_dict = {}
        for sym in corr.columns:
            corr_dict[sym] = {
                other: round(float(corr.loc[sym, other]), 4)
                for other in corr.columns
            }

        # Summary stats
        stats = {}
        for sym in closes.columns:
            s = closes[sym].dropna()
            r = rets[sym].dropna() if sym in rets.columns else pd.Series(dtype=float)
            if s.empty:
                continue
            stats[sym] = {
                "return_total": round((s.iloc[-1]/s.iloc[0] - 1)*100, 2),
                "volatility":   round(float(r.std() * np.sqrt(252) * 100), 2) if len(r) > 1 else None,
                "sharpe":       _calc_sharpe(r),
                "max_drawdown": round(_max_drawdown(s) * 100, 2),
            }

        return {
            "success":    True,
            "symbols":    list(closes.columns),
            "period":     period,
            "interval":   interval,
            "corr_matrix": corr_dict,
            "stats":       stats,
            "data_points": len(rets),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── 多资产组合回测 ─────────────────────────────────────────────────────────────

def portfolio_backtest(params: dict) -> dict:
    """
    多资产组合历史回测。

    参数：
      symbols:   list[str]，如 ["AAPL","MSFT","GOOG"]
      weights:   list[float]，权重（自动归一化）；为空则等权
      period:    "1y" / "2y" / "3y" / "5y"（默认 "2y"）
      benchmark: 基准标的（默认 "SPY"）
      rebalance: 再平衡频率 ("monthly" / "quarterly" / "none"，默认 "monthly")
    """
    if not _HAS_YF or not _HAS_PD or not _HAS_NP:
        return {"success": False, "error": "需要 yfinance + pandas + numpy"}

    symbols   = params.get("symbols", [])
    weights_r = params.get("weights", [])
    period    = str(params.get("period", "2y"))
    benchmark = str(params.get("benchmark", "SPY"))
    rebalance = str(params.get("rebalance", "monthly"))

    if not symbols:
        return {"success": False, "error": "symbols 不能为空"}
    if len(symbols) > 15:
        symbols = symbols[:15]

    # Normalize weights
    if weights_r and len(weights_r) == len(symbols):
        ws = [float(w) for w in weights_r]
        total = sum(ws)
        weights = [w / total for w in ws]
    else:
        weights = [1 / len(symbols)] * len(symbols)

    all_syms = list(dict.fromkeys(symbols + [benchmark]))
    try:
        raw = yf.download(all_syms, period=period, progress=False, auto_adjust=True)
        if raw.empty:
            return {"success": False, "error": "数据下载失败"}

        closes = raw["Close"] if isinstance(raw.columns, pd.MultiIndex) else raw
        closes = closes.dropna(how="all", axis=1)

        # Align symbols with available data
        avail = [s for s in symbols if s in closes.columns]
        if not avail:
            return {"success": False, "error": "所有标的数据均不可用"}
        if len(avail) < len(symbols):
            missing = [s for s in symbols if s not in closes.columns]
            weights = [weights[i] for i, s in enumerate(symbols) if s in closes.columns]
            wt_total = sum(weights)
            weights = [w / wt_total for w in weights]
            symbols = avail

        port_data = closes[symbols].ffill().dropna()
        if benchmark in closes.columns:
            bench_data = closes[benchmark].ffill().dropna()
        else:
            bench_data = None

        # Rebalance
        if rebalance == "monthly":
            freq = "MS"
        elif rebalance == "quarterly":
            freq = "QS"
        else:
            freq = None

        # Compute portfolio returns
        if freq and len(port_data) > 0:
            port_rets = port_data.pct_change().dropna()
            # Monthly rebalancing: compound weighted returns per period
            port_monthly_rets = (port_rets + 1).resample(freq).prod() - 1
            port_rets_w = (port_monthly_rets[symbols] * weights).sum(axis=1)
            # Expand back to daily for metrics
            port_cum = (1 + port_rets_w).cumprod()
        else:
            port_rets = port_data.pct_change().dropna()
            port_rets_w = (port_rets[symbols] * weights).sum(axis=1)
            port_cum = (1 + port_rets_w).cumprod()

        total_return = float(port_cum.iloc[-1] - 1) * 100
        annual_vol   = float(port_rets_w.std() * np.sqrt(252 if freq else 252) * 100)
        sharpe       = _calc_sharpe(port_rets_w)
        max_dd       = round(_max_drawdown(port_cum) * 100, 2)

        # Benchmark
        bench_stats = None
        if bench_data is not None:
            bench_rets = bench_data.pct_change().dropna()
            b_cum = (1 + bench_rets).cumprod()
            bench_stats = {
                "symbol":     benchmark,
                "total_return": round((float(b_cum.iloc[-1]) - 1)*100, 2),
                "volatility": round(float(bench_rets.std() * np.sqrt(252) * 100), 2),
                "sharpe":     _calc_sharpe(bench_rets),
                "max_drawdown": round(_max_drawdown(b_cum) * 100, 2),
            }

        # Allocation breakdown
        allocation = [{"symbol": s, "weight_pct": round(w*100, 1),
                       "weight_abs": w} for s, w in zip(symbols, weights)]

        # Performance curve (monthly points)
        if freq:
            curve_pts = [{"date": str(d.date()), "value": round(float(v), 4)}
                         for d, v in port_cum.items()]
        else:
            curve_pts = [{"date": str(d.date()), "value": round(float(v), 4)}
                         for d, v in port_cum.iloc[::5].items()]  # every 5 days

        return {
            "success":     True,
            "symbols":     symbols,
            "weights":     weights,
            "allocation":  allocation,
            "period":      period,
            "rebalance":   rebalance,
            "portfolio": {
                "total_return_pct": round(total_return, 2),
                "annual_vol_pct":   round(annual_vol, 2),
                "sharpe_ratio":     sharpe,
                "max_drawdown_pct": max_dd,
                "calmar_ratio":     round(total_return / abs(max_dd), 2) if max_dd != 0 else None,
            },
            "benchmark":   bench_stats,
            "curve":       curve_pts[-60:],  # last 60 points for chart
            "data_points": len(port_rets_w),
        }
    except Exception as e:
        logger.exception("Portfolio backtest error")
        return {"success": False, "error": str(e)}


# ── 自定义因子 DSL ─────────────────────────────────────────────────────────────

def eval_custom_factor(params: dict) -> dict:
    """
    计算自定义因子表达式。

    参数：
      symbol:  标的代码
      period:  历史区间（默认 "1y"）
      expr:    因子表达式字符串（Python 数学语法，可使用变量）
               可用变量：close, open, high, low, volume, returns
               可用函数：sma(n), ema(n), std(n), rsi(n), atr(n)
               示例：
                 "(close - sma(20)) / std(20)"     → Z-Score
                 "sma(5) / sma(20) - 1"            → 短期动量
                 "volume / sma(20, volume)"         → 量比
    """
    if not _HAS_YF or not _HAS_PD or not _HAS_NP:
        return {"success": False, "error": "需要 yfinance + pandas + numpy"}

    symbol = str(params.get("symbol", "")).upper()
    period = str(params.get("period", "1y"))
    expr   = str(params.get("expr", "")).strip()

    if not symbol or not expr:
        return {"success": False, "error": "symbol 和 expr 均为必填"}

    # Security: block imports and exec
    forbidden = ["import", "__", "exec(", "eval(", "open(", "os.", "sys."]
    for tok in forbidden:
        if tok in expr:
            return {"success": False, "error": f"表达式包含禁止关键字: {tok}"}

    try:
        tkr = yf.Ticker(symbol)
        hist = tkr.history(period=period, auto_adjust=True)
        if hist.empty:
            return {"success": False, "error": f"无法获取 {symbol} 历史数据"}

        close  = hist["Close"].astype(float)
        open_  = hist["Open"].astype(float)
        high   = hist["High"].astype(float)
        low    = hist["Low"].astype(float)
        volume = hist["Volume"].astype(float)
        returns = close.pct_change()

        def sma(n, series=None):
            s = series if series is not None else close
            return s.rolling(n).mean()

        def ema(n, series=None):
            s = series if series is not None else close
            return s.ewm(span=n, adjust=False).mean()

        def std(n, series=None):
            s = series if series is not None else close
            return s.rolling(n).std()

        def rsi(n=14):
            delta = close.diff()
            gain  = delta.clip(lower=0).rolling(n).mean()
            loss  = (-delta.clip(upper=0)).rolling(n).mean()
            rs    = gain / loss
            return 100 - (100 / (1 + rs))

        def atr(n=14):
            tr = pd.concat([
                high - low,
                (high - close.shift()).abs(),
                (low  - close.shift()).abs(),
            ], axis=1).max(axis=1)
            return tr.rolling(n).mean()

        ns = dict(
            close=close, open=open_, high=high, low=low,
            volume=volume, returns=returns,
            sma=sma, ema=ema, std=std, rsi=rsi, atr=atr,
            pd=pd, np=np, abs=abs, min=min, max=max,
        )

        result_series = eval(expr, {"__builtins__": {}}, ns)  # noqa: S307
        if hasattr(result_series, "dropna"):
            result_series = result_series.dropna()

        latest_val = float(result_series.iloc[-1]) if len(result_series) > 0 else None

        # Last 20 values for chart
        tail = result_series.tail(20)
        series_pts = [{"date": str(d.date()), "value": round(float(v), 6)}
                      for d, v in tail.items()]

        return {
            "success":   True,
            "symbol":    symbol,
            "expr":      expr,
            "latest":    latest_val,
            "series":    series_pts,
            "data_len":  len(result_series),
        }
    except Exception as e:
        return {"success": False, "error": f"表达式计算失败: {e}"}


# ── 通用数据导入 ───────────────────────────────────────────────────────────────

def load_csv_data(params: dict) -> dict:
    """
    从 CSV 文件加载数据到内存 DuckDB 表。

    参数：
      path:       CSV 文件路径
      table_name: 目标表名（默认由文件名推导）
      encoding:   编码（默认 utf-8）
    """
    if not _HAS_PD:
        return {"success": False, "error": "pandas 未安装"}

    path = Path(str(params.get("path", ""))).expanduser()
    if not path.exists():
        return {"success": False, "error": f"文件不存在: {path}"}
    if path.suffix.lower() not in (".csv", ".tsv", ".txt"):
        return {"success": False, "error": "仅支持 CSV/TSV 文件"}

    enc   = params.get("encoding", "utf-8")
    tname = str(params.get("table_name", path.stem))

    try:
        df = pd.read_csv(str(path), encoding=enc)
        if _HAS_DUCK:
            conn = _get_duck_conn()
            conn.register(tname, df)
            _duck_tables[tname] = True

        return {
            "success":    True,
            "table_name": tname,
            "rows":       len(df),
            "columns":    list(df.columns),
            "sample":     df.head(3).to_dict("records"),
            "path":       str(path),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# ── Private Helpers ───────────────────────────────────────────────────────────

def _calc_sharpe(rets, rf_daily: float = 0.0) -> Optional[float]:
    if not _HAS_NP or not hasattr(rets, "__len__") or len(rets) < 2:
        return None
    try:
        excess = rets - rf_daily
        if float(excess.std()) == 0:
            return None
        return round(float(excess.mean() / excess.std() * np.sqrt(252)), 3)
    except Exception:
        return None


def _max_drawdown(cum_series) -> float:
    if not _HAS_NP or not _HAS_PD:
        return 0.0
    try:
        s = pd.Series(cum_series)
        roll_max = s.cummax()
        drawdown = (s - roll_max) / roll_max
        return float(drawdown.min())
    except Exception:
        return 0.0
